import reflex as rx
from io import BytesIO
from datetime import datetime
import json
import logging
from openpyxl import Workbook, load_workbook
from app.states.order_state import OrderState
from app.states.menu_state import MenuState
from app.states.inventory_state import InventoryState
from app.states.staff_state import StaffState
from app.states.customer_state import CustomerState
from app.states.delivery_state import DeliveryState
from app.states.billing_state import BillingState

UPLOAD_ID = "excel_upload"

SHEET_CONFIG: list[tuple[str, list[str]]] = [
    (
        "orders",
        ["id", "customer", "items", "status", "channel", "time", "total"],
    ),
    (
        "menu",
        [
            "id",
            "name",
            "description",
            "price",
            "category",
            "modifiers",
            "available",
            "seasonal_start",
            "seasonal_end",
            "image",
        ],
    ),
    (
        "inventory",
        [
            "id",
            "name",
            "quantity",
            "unit",
            "reorder_level",
            "supplier",
            "status",
        ],
    ),
    ("staff", ["id", "name", "role", "email", "phone", "status"]),
    (
        "customers",
        [
            "id",
            "name",
            "email",
            "phone",
            "total_orders",
            "total_spent",
            "last_order",
            "is_vip",
        ],
    ),
    (
        "deliveries",
        ["order_id", "customer", "driver", "status", "eta", "platform"],
    ),
    ("billing", ["id", "customer", "amount", "status", "date", "method"]),
]


def _safe_cell(v) -> str | int | float | bool:
    if v is None:
        return ""
    if isinstance(v, (int, float, bool, str)):
        return v
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    return str(v)


class ExcelState(rx.State):
    status_message: str = ""
    is_error: bool = False
    last_export_time: str = ""
    last_import_time: str = ""
    is_processing: bool = False

    @rx.event
    async def export_data(self):
        self.is_processing = True
        self.status_message = ""
        self.is_error = False
        try:
            order_state = await self.get_state(OrderState)
            menu_state = await self.get_state(MenuState)
            inventory_state = await self.get_state(InventoryState)
            staff_state = await self.get_state(StaffState)
            customer_state = await self.get_state(CustomerState)
            delivery_state = await self.get_state(DeliveryState)
            billing_state = await self.get_state(BillingState)

            data_map: dict[str, list[dict]] = {
                "orders": list(order_state.orders),
                "menu": list(menu_state.items),
                "inventory": list(inventory_state.items),
                "staff": list(staff_state.staff),
                "customers": list(customer_state.customers),
                "deliveries": list(delivery_state.deliveries),
                "billing": list(billing_state.invoices),
            }

            wb = Workbook()
            wb.remove(wb.active)
            for sheet_name, headers in SHEET_CONFIG:
                ws = wb.create_sheet(title=sheet_name)
                ws.append(headers)
                for row in data_map.get(sheet_name, []):
                    ws.append([_safe_cell(row.get(h)) for h in headers])

            stream = BytesIO()
            wb.save(stream)
            stream.seek(0)
            data_bytes = stream.read()

            self.last_export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.status_message = "Workbook exported successfully."
            self.is_error = False
            self.is_processing = False
            yield rx.toast("Excel workbook exported")
            yield rx.download(
                data=data_bytes, filename="cloud_kitchen_data.xlsx"
            )
        except Exception as e:
            logging.exception(f"Export failed: {e}")
            self.is_error = True
            self.status_message = f"Export failed: {e}"
            self.is_processing = False

    @rx.event
    async def handle_upload(self, files: list[rx.UploadFile]):
        if not files:
            self.is_error = True
            self.status_message = "No file selected."
            return
        self.is_processing = True
        self.status_message = ""
        self.is_error = False
        try:
            file = files[0]
            data = await file.read()
            if not file.name.lower().endswith(".xlsx"):
                self.is_error = True
                self.status_message = (
                    "Invalid file type. Please upload a .xlsx file."
                )
                self.is_processing = False
                return

            wb = load_workbook(BytesIO(data), data_only=True)
            available = set(wb.sheetnames)
            expected = {name for name, _ in SHEET_CONFIG}
            missing = expected - available

            imported_counts: dict[str, int] = {}

            for sheet_name, headers in SHEET_CONFIG:
                if sheet_name not in available:
                    continue
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    imported_counts[sheet_name] = 0
                    continue
                file_headers = [
                    str(h) if h is not None else "" for h in rows[0]
                ]
                records: list[dict] = []
                for raw in rows[1:]:
                    if raw is None or all(v is None for v in raw):
                        continue
                    record: dict = {}
                    for idx, header in enumerate(file_headers):
                        if header not in headers:
                            continue
                        value = raw[idx] if idx < len(raw) else None
                        record[header] = self._coerce(sheet_name, header, value)
                    for h in headers:
                        if h not in record:
                            record[h] = self._default_for(sheet_name, h)
                    records.append(record)
                imported_counts[sheet_name] = len(records)
                await self._apply_records(sheet_name, records)

            total = sum(imported_counts.values())
            self.last_import_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if missing:
                self.status_message = (
                    f"Imported {total} records across {len(imported_counts)} sheet(s). "
                    f"Missing sheets skipped: {', '.join(sorted(missing))}."
                )
                self.is_error = False
            else:
                self.status_message = f"Imported {total} records across all {len(imported_counts)} sheets."
                self.is_error = False
            self.is_processing = False
            yield rx.toast("Import complete")
        except Exception as e:
            logging.exception(f"Import failed: {e}")
            self.is_error = True
            self.status_message = f"Import failed: {e}"
            self.is_processing = False

    def _coerce(self, sheet: str, key: str, value):
        if value is None:
            if key in {
                "price",
                "total",
                "amount",
                "quantity",
                "reorder_level",
                "total_spent",
            }:
                return 0.0
            if key in {"total_orders"}:
                return 0
            if key in {"available", "is_vip"}:
                return False
            if key == "modifiers":
                return []
            return ""
        if key == "modifiers":
            if isinstance(value, list):
                return value
            return [s.strip() for s in str(value).split(",") if s.strip()]
        if key in {"available", "is_vip"}:
            if isinstance(value, bool):
                return value
            return str(value).strip().lower() in {"true", "1", "yes"}
        if key in {
            "price",
            "total",
            "amount",
            "quantity",
            "reorder_level",
            "total_spent",
        }:
            try:
                return float(value)
            except Exception:
                logging.exception("Unexpected error")
                return 0.0
        if key == "total_orders":
            try:
                return int(float(value))
            except Exception:
                logging.exception("Unexpected error")
                return 0
        return str(value)

    def _default_for(self, sheet: str, key: str):
        if key == "modifiers":
            return []
        if key in {"available", "is_vip"}:
            return False
        if key in {
            "price",
            "total",
            "amount",
            "quantity",
            "reorder_level",
            "total_spent",
        }:
            return 0.0
        if key == "total_orders":
            return 0
        return ""

    async def _apply_records(self, sheet: str, records: list[dict]):
        if sheet == "orders":
            s = await self.get_state(OrderState)
            s.orders = records
            s.orders_storage = json.dumps(records)
        elif sheet == "menu":
            s = await self.get_state(MenuState)
            s.items = records
            s.menu_storage = json.dumps(records)
        elif sheet == "inventory":
            s = await self.get_state(InventoryState)
            s.items = records
            s.inventory_storage = json.dumps(records)
        elif sheet == "staff":
            s = await self.get_state(StaffState)
            s.staff = records
            s.staff_storage = json.dumps(records)
        elif sheet == "customers":
            s = await self.get_state(CustomerState)
            s.customers = records
            s.customer_storage = json.dumps(records)
        elif sheet == "deliveries":
            s = await self.get_state(DeliveryState)
            s.deliveries = records
            s.delivery_storage = json.dumps(records)
        elif sheet == "billing":
            s = await self.get_state(BillingState)
            s.invoices = records
            s.billing_storage = json.dumps(records)